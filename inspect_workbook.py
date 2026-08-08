import os
import glob
import openpyxl

folder = r"e:\Research\NLDC Dataset\01. Carbon Intensity\Raw Data"
files = sorted(glob.glob(os.path.join(folder, '*.xlsx')))
print('files', len(files))
if not files:
    raise SystemExit('No xlsx files found')
p = files[0]
wb = openpyxl.load_workbook(p, data_only=True)
print('file', os.path.basename(p))
print('sheets', wb.sheetnames)
ws = wb.active
print('rows', ws.max_row, 'cols', ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
    print(row)
