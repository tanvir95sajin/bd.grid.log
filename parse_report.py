import json
import re
from pathlib import Path

import openpyxl
from openpyxl import load_workbook


def _cell_value(row, index):
    if index < len(row):
        return row[index]
    return None


def _find_row(ws, value):
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(cell == value for cell in row if cell is not None):
            return row_idx
    return None


def _extract_date(ws, path):
    filename_match = re.search(r"(\d{1,2})[-\s_](\d{1,2})[-\s_](\d{4})", path.name)
    if filename_match:
        day, month, year = filename_match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"

    for row in ws.iter_rows(values_only=True):
        for idx, cell in enumerate(row):
            if cell == "Date :":
                date_cell = _cell_value(row, idx + 1)
                if isinstance(date_cell, str):
                    m = re.search(r"(\d{2})-(\d{2})-(\d{4})", date_cell)
                    if m:
                        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"

    return None


def _extract_summary(ws):
    mapping = {
        "Day Peak Generation": "day_peak_generation",
        "Day Peak Demand": "day_peak_demand",
        "Evening Peak Generation": "evening_peak_generation",
        "Evening Peak Demand": "evening_peak_demand",
        "Minimum Generation of the Day": "minimum_generation_of_the_day",
        "Maximum Generation of the Day": "maximum_generation_of_the_day",
    }
    result = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        first = next((cell for cell in row if cell is not None), None)
        if isinstance(first, str) and first in mapping:
            value = None
            for cell in row:
                if isinstance(cell, (int, float)):
                    value = cell
                    break
            if value is not None:
                result[mapping[first]] = value
    return result


def _extract_zone_summary(ws):
    rows = []
    zone_headers = None
    for row in ws.iter_rows(values_only=True):
        if not any(cell is not None for cell in row):
            continue
        if any(isinstance(cell, str) and "Zone-wise Generation Summary" in str(cell) for cell in row):
            continue
        first_cell = next((cell for cell in row if cell is not None), None)
        if isinstance(first_cell, str) and first_cell.endswith("Zone"):
            zone_name = first_cell
            values = []
            for cell in row:
                if isinstance(cell, (int, float)):
                    values.append(cell)
            if values:
                rows.append({"zone": zone_name, "values": values})
    return rows


def _load_workbook(path):
    try:
        return load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return load_workbook(path, data_only=True)


def parse_one(path, output_dir=None):
    path = Path(path)
    wb = _load_workbook(path)
    ws = wb.active
    date_str = _extract_date(ws, path)
    if not date_str:
        raise ValueError(f"Could not find a date in {path.name}")

    summary = _extract_summary(ws)
    zones = _extract_zone_summary(ws)
    data = {
        "date": date_str,
        "summary": summary,
        "zones": zones,
    }

    if output_dir is None:
        output_path = path.with_suffix('.json')
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{date_str}.json"

    with output_path.open('w', encoding='utf-8') as fh:
        json.dump(data, fh, indent=2)
    return date_str
