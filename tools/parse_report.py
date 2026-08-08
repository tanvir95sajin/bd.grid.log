"""
Parse a PGCB/NLDC daily report (.xlsx) into a single JSON file for the
Grid Log website.

Usage:
    python3 parse_report.py path/to/Daily_Report_DD-MM-YYYY.xlsx

Writes: data/YYYY-MM-DD.json  (dated by the ACTUAL data date, which is the
day BEFORE the report date in the filename - PGCB's daily report always
covers "yesterday" in full plus a forecast for today).
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl

# Approximate CO2 emission factors, kg CO2 per kWh generated.
# These are rough, widely-used planning figures, not audited emissions data.
EMISSION_FACTORS = {
    "Coal": 0.95,
    "Gas-Public": 0.45,
    "Gas-Pvt": 0.45,
    "HFO-Public": 0.73,
    "HFO-Pvt": 0.73,
    "Adani": 0.95,       # imported power from a coal-fired plant across the border
    "Solar": 0.0,
    "Hydro": 0.0,
    "HVDC": None,        # cross-border import, source mix not published - not estimated
    "Nepal": None,
    "Tripura": None,
}

FUEL_LABELS = {
    "Gas-Public": "Gas (public)",
    "Gas-Pvt": "Gas (private)",
    "HVDC": "Import (HVDC)",
    "Nepal": "Import (Nepal)",
    "Tripura": "Import (Tripura)",
    "Adani": "Import (Adani, coal)",
    "Hydro": "Hydro",
    "Coal": "Coal",
    "Solar": "Solar",
    "HFO-Public": "Oil - HFO (public)",
    "HFO-Pvt": "Oil - HFO (private)",
}


def to_num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 3)
    return None


def norm_name(s):
    return " ".join(str(s).split()).strip()


def base_name(name):
    """Strip a trailing (Gas)/(HSD)/(HFO) tag some sheets split a unit by,
    e.g. GenLog listing 'Sikalbaha Peaking GT (Gas)' and '... (HSD)' as two
    columns for what is really one physical, dual-fired unit."""
    m = re.match(r"^(.*)\s*\((Gas|HSD|HFO)\)\s*$", name.strip())
    return norm_name(m.group(1)) if m else name


def is_header_repeat(sl, name):
    """PGCB's print layout repeats the column headers as a data row every
    time the sheet breaks to a new printed page. Catch and skip those."""
    name_l = str(name).lower()
    if "name of the" in name_l and "power station" in name_l:
        return True
    if str(sl).strip().lower() in ("sl.", "sl"):
        return True
    return False


def parse_capacity_str(s):
    """'1*260' or '2*150+1*150' -> total MW as a float."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return round(float(s), 2)
    total = 0.0
    found = False
    for part in str(s).split("+"):
        part = part.strip()
        m = re.match(r"^([\d.]+)\s*\*\s*([\d.]+)$", part)
        if m:
            total += float(m.group(1)) * float(m.group(2))
            found = True
        else:
            try:
                total += float(part)
                found = True
            except ValueError:
                pass
    return round(total, 2) if found else None


def parse_p1(ws):
    out = {}
    grid = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                grid[(r, c)] = v

    def find_label(label_substr):
        for (r, c), v in grid.items():
            if isinstance(v, str) and label_substr.lower() in v.lower():
                return r, c
        return None

    def value_right_of(label_substr, offset=3):
        pos = find_label(label_substr)
        if not pos:
            return None
        r, c = pos
        return to_num(grid.get((r, c + offset)))

    out["day_peak_generation_mw"] = value_right_of("Day Peak Generation")
    out["day_peak_demand_mw"] = value_right_of("Day Peak Demand")
    out["evening_peak_generation_mw"] = value_right_of("Evening Peak Generation")
    out["evening_peak_demand_mw"] = value_right_of("Evening Peak Demand")
    out["min_generation_mw"] = value_right_of("Minimum Generation")
    out["max_generation_mw"] = value_right_of("Maximum Generation")
    out["energy_generated_gwh"] = value_right_of("Energy Generated")
    out["energy_unserved_gwh"] = value_right_of("Energy Unserverd") or value_right_of("Energy Unserved")
    out["energy_demand_gwh"] = value_right_of("Energy Demand")
    out["max_temperature_c"] = value_right_of("Maximum Temperature")
    out["total_gas_supplied_mmcfd"] = value_right_of("Total Gas Supplied")
    return out


def parse_zone_fuel_summary(ws):
    """'Zone-wise Generation Summary (MKWHr.)' block, rows 17-25 of P1."""
    fuels = ["Gas", "Coal", "HFO", "HSD", "Hydro", "Solar", "Import"]
    zones = {}
    for r in range(17, 26):
        name = ws.cell(row=r, column=2).value
        if not name or "zone" not in str(name).lower():
            continue
        zone = norm_name(name).replace(" Zone", "")
        vals = {}
        for i, fuel in enumerate(fuels):
            v = to_num(ws.cell(row=r, column=5 + i).value)
            vals[fuel] = v or 0
        zones[zone] = vals
    return zones


def parse_p4(ws):
    hourly = []
    for r in range(9, ws.max_row + 1):
        hour = ws.cell(row=r, column=2).value
        gen = to_num(ws.cell(row=r, column=3).value)
        shed = to_num(ws.cell(row=r, column=5).value)
        if hour is None or gen is None:
            continue
        hour_str = hour.strftime("%H:%M") if hasattr(hour, "strftime") else str(hour)
        demand = round(gen + (shed or 0), 3)
        hourly.append({"hour": hour_str, "generation_mw": gen, "load_shed_mw": shed or 0, "demand_mw": demand})
    return hourly


def parse_p3_divisions(ws):
    zones = []
    labels = [
        (9, 5, "Dhaka"), (10, 5, "Chattogram"), (11, 5, "Cumilla"),
        (12, 5, "Mymensingh"), (13, 5, "Sylhet"),
        (9, 12, "Khulna"), (10, 12, "Barishal"), (11, 12, "Rajshahi"), (12, 12, "Rangpur"),
    ]
    for r, c, name in labels:
        val = to_num(ws.cell(row=r, column=c).value)
        if val is not None:
            zones.append({"division": name, "evening_peak_load_mw": val})
    return sorted(zones, key=lambda z: -z["evening_peak_load_mw"])


def time_to_str(v):
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%H:%M")
    return str(v)


def parse_p3_substations(ws):
    """Three side-by-side column blocks, rows 17-84: Sl No | Sub-station | Load (MW) | Time."""
    subs = {}
    for name_col in (3, 7, 11):
        load_col, time_col = name_col + 1, name_col + 2
        for r in range(17, 85):
            name = ws.cell(row=r, column=name_col).value
            load = to_num(ws.cell(row=r, column=load_col).value)
            if not name or load is None:
                continue
            key = norm_name(name)
            subs[key] = {"name": key, "load_mw": load, "load_time": time_to_str(ws.cell(row=r, column=time_col).value)}
    return subs


def parse_voltage(ws):
    """Two side-by-side column blocks, rows 7-159, with kV section headers
    ('400 kV', '230 kV', '132 kV') interleaved as rows."""
    subs = {}
    for name_col in (1, 6):
        max_col, maxt_col, min_col, mint_col = name_col + 1, name_col + 2, name_col + 3, name_col + 4
        level = None
        for r in range(7, 160):
            name = ws.cell(row=r, column=name_col).value
            if isinstance(name, str) and "kv" in name.lower() and re.match(r"^\d+\s*kv$", name.strip().lower()):
                level = name.strip()
                continue
            maxv = to_num(ws.cell(row=r, column=max_col).value)
            if not name or maxv is None:
                continue
            key = norm_name(name)
            subs[key] = {
                "name": key,
                "voltage_level": level,
                "max_voltage_kv": maxv,
                "max_voltage_time": time_to_str(ws.cell(row=r, column=maxt_col).value),
                "min_voltage_kv": to_num(ws.cell(row=r, column=min_col).value),
                "min_voltage_time": time_to_str(ws.cell(row=r, column=mint_col).value),
            }
    return subs


def parse_en_curve(ws):
    fuels = [ws.cell(row=3, column=c).value for c in range(2, 13)]
    intervals = []
    sums = [0.0] * len(fuels)
    n = 0
    for r in range(4, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        vals = [ws.cell(row=r, column=c).value for c in range(2, 13)]
        if t is None or not all(isinstance(v, (int, float)) for v in vals):
            continue
        t_str = t.strftime("%H:%M") if hasattr(t, "strftime") else str(t)
        row_obj = {"time": t_str}
        for f, v in zip(fuels, vals):
            row_obj[f] = round(v, 2)
            sums[fuels.index(f)] += v
        intervals.append(row_obj)
        n += 1

    avg_mw = {f: round(s / n, 1) for f, s in zip(fuels, sums)} if n else {}
    total_avg = sum(avg_mw.values()) if avg_mw else 1
    energy_mwh = {f: round(avg_mw[f] * 24, 1) for f in avg_mw}

    co2_tonnes = {}
    co2_total = 0.0
    co2_unestimated_mwh = 0.0
    for f, mwh in energy_mwh.items():
        factor = EMISSION_FACTORS.get(f)
        if factor is None:
            co2_unestimated_mwh += mwh
            continue
        tonnes = round(mwh * 1000 * factor / 1000, 1)
        co2_tonnes[f] = tonnes
        co2_total += tonnes

    fuel_mix_pct = {f: round(v / total_avg * 100, 1) for f, v in avg_mw.items()}

    return {
        "half_hourly": intervals,
        "avg_mw_by_fuel": avg_mw,
        "energy_mwh_by_fuel": energy_mwh,
        "fuel_mix_pct": fuel_mix_pct,
        "co2_tonnes_by_fuel": co2_tonnes,
        "co2_total_tonnes_estimated": round(co2_total, 1),
        "co2_unestimated_mwh": round(co2_unestimated_mwh, 1),
        "fuel_labels": FUEL_LABELS,
    }


def parse_genlog_columns(ws):
    """Returns {column_name: {hourly_mw, peak_mw, energy_kwh}} for every
    named column in GenLog, keyed EXACTLY as GenLog names it (this can
    include fuel-mode-split names like 'X (Gas)' / 'X (HSD)')."""
    headers = [ws.cell(row=11, column=c).value for c in range(1, ws.max_column + 1)]
    plant_cols = []
    for i, h in enumerate(headers):
        if h and h not in ("Hour", "Eastern Total", "Western Total", "National Grid Total", "Water Level"):
            plant_cols.append((i + 1, norm_name(h)))

    hour_rows = list(range(13, 38))  # 00:00 .. 23:00 (plus the 19:30 half-hour row)
    out = {}
    for col, name in plant_cols:
        series = []
        for r in hour_rows:
            hour = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=col).value
            hour_str = hour.strftime("%H:%M") if hasattr(hour, "strftime") else str(hour)
            series.append({"hour": hour_str, "mw": to_num(val) or 0})
        total_kwh = None
        for r in range(36, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if isinstance(label, str) and "total kwh" in label.lower():
                total_kwh = to_num(ws.cell(row=r, column=col).value)
                break
        out[name] = {"hourly_mw": series, "energy_kwh": total_kwh}
    return out


def parse_p2(ws):
    """Keyed by plant name -> producer/capacity/remarks. Includes every
    named row regardless of whether it has an official Sl. number, since
    this is used purely as a lookup table for remarks etc."""
    out = {}
    for r in range(11, ws.max_row + 1):
        name = ws.cell(row=r, column=3).value
        sl = ws.cell(row=r, column=2).value
        if not name or is_header_repeat(sl, name):
            continue
        key = norm_name(name)
        out[key] = {
            "producer": ws.cell(row=r, column=5).value,
            "present_capacity_mw": to_num(ws.cell(row=r, column=7).value),
            "peak_hour_generation_mw": to_num(ws.cell(row=r, column=8).value),
            "energy_generated_kwh": to_num(ws.cell(row=r, column=9).value),
            "remarks": ws.cell(row=r, column=10).value,
        }
    return out


def parse_official_plants(ws):
    """The Forecast sheet's Sl.-numbered rows ARE PGCB's own plant
    registry for the day - this is the authoritative plant list and count
    (143 as of mid-2026), not a name-union across sheets."""
    out = {}
    for r in range(11, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        sl = ws.cell(row=r, column=1).value
        if not name or is_header_repeat(sl, name):
            continue
        if sl in (None, ""):
            continue  # no official Sl number -> not a separately counted plant
        key = norm_name(name)
        cap_str = ws.cell(row=r, column=5).value
        out[key] = {
            "sl": str(sl).strip(),
            "fuel": ws.cell(row=r, column=3).value,
            "producer_forecast": ws.cell(row=r, column=4).value,
            "installed_capacity_mw": parse_capacity_str(cap_str),
            "present_capacity_mw_forecast": to_num(ws.cell(row=r, column=6).value),
        }
    return out


def genlog_hourly_for(plant_name, genlog_cols):
    """Sum every GenLog column that belongs to this plant - handles the
    case where GenLog splits one dual-fired unit into '<name> (Gas)' and
    '<name> (HSD)' columns for the same physical, Sl.-numbered plant."""
    matches = [v for k, v in genlog_cols.items() if k == plant_name or base_name(k) == plant_name]
    if not matches:
        return None, 0, None

    hours = [h["hour"] for h in matches[0]["hourly_mw"]]
    summed = []
    for i, hour in enumerate(hours):
        summed.append({"hour": hour, "mw": round(sum(m["hourly_mw"][i]["mw"] for m in matches), 2)})
    peak = max((h["mw"] for h in summed), default=0)
    energy = sum((m["energy_kwh"] or 0) for m in matches) or None
    return summed, peak, energy


def build_plants(official, genlog_cols, p2):
    plants = {}
    for name, off in official.items():
        p = p2.get(name)
        hourly, peak_mw, energy_kwh_genlog = genlog_hourly_for(name, genlog_cols)
        energy_kwh = energy_kwh_genlog or (p or {}).get("energy_generated_kwh") or 0
        status = "running" if (peak_mw > 0 or energy_kwh > 0) else "offline"
        cause = (p or {}).get("remarks") if status == "offline" else None
        if cause in ("-", "", None):
            cause = None if status == "running" else "Not stated in report"

        cap_for_utilization = (p or {}).get("present_capacity_mw") or off.get("present_capacity_mw_forecast") or off.get("installed_capacity_mw")
        utilization_pct = round(peak_mw / cap_for_utilization * 100, 1) if cap_for_utilization else None

        plants[name] = {
            "sl": off["sl"],
            "fuel": off["fuel"],
            "producer": (p or {}).get("producer") or off.get("producer_forecast"),
            "installed_capacity_mw": off.get("installed_capacity_mw"),
            "present_capacity_mw": (p or {}).get("present_capacity_mw") or off.get("present_capacity_mw_forecast"),
            "utilization_pct": utilization_pct,
            "status": status,
            "cause": cause,
            "peak_mw": peak_mw,
            "energy_kwh": energy_kwh,
            "hourly_mw": hourly,
        }
    return plants


def parse_one(path: Path) -> str:
    """Parse a single .xlsx report and write its dated JSON file.
    Returns the data date string (YYYY-MM-DD) written."""
    wb = openpyxl.load_workbook(path, data_only=True)

    m = re.search(r"(\d{2})-(\d{2})-(\d{4})", path.stem)
    if m:
        dd, mm, yyyy = m.groups()
        issue_date = datetime(int(yyyy), int(mm), int(dd))
        data_date = issue_date - timedelta(days=1)
    else:
        data_date = datetime.today()
    date_str = data_date.strftime("%Y-%m-%d")

    official = parse_official_plants(wb["Forecast"])
    genlog_cols = parse_genlog_columns(wb["GenLog"])
    p2 = parse_p2(wb["P2"])
    plants = build_plants(official, genlog_cols, p2)

    p3_load = parse_p3_substations(wb["P3"])
    voltage = parse_voltage(wb["Voltage"])
    all_sub_names = set(p3_load) | set(voltage)
    substations = {}
    for name in all_sub_names:
        entry = {"name": name}
        entry.update(p3_load.get(name, {}))
        entry.update({k: v for k, v in voltage.get(name, {}).items() if k != "name"})
        substations[name] = entry

    result = {
        "date": date_str,
        "source_file": path.name,
        "summary": parse_p1(wb["P1"]),
        "hourly": parse_p4(wb["P4"]),
        "divisions": parse_p3_divisions(wb["P3"]),
        "zone_fuel_summary_mkwhr": parse_zone_fuel_summary(wb["P1"]),
        "fuel_mix": parse_en_curve(wb["En-Curve"]),
        "plants": plants,
        "substations": substations,
    }

    out_dir = Path(__file__).resolve().parent.parent / "data"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"{date_str}.json"
    with open(out_path, "w") as f:
        json.dump(result, f, indent=1)

    running = sum(1 for p in plants.values() if p["status"] == "running")
    sl_numbers = sorted((int(float(p["sl"])) for p in plants.values() if p["sl"].replace(".", "", 1).isdigit()))
    print(f"Wrote {out_path}")
    print(f"  {len(plants)} officially Sl.-numbered plants ({running} running, {len(plants) - running} offline)")
    if sl_numbers:
        gaps = sorted(set(range(sl_numbers[0], sl_numbers[-1] + 1)) - set(sl_numbers))
        print(f"  Sl. numbers span {sl_numbers[0]}-{sl_numbers[-1]}; missing from this file: {gaps if gaps else 'none'}")
    print(f"  {len(substations)} substations")
    print(f"  date={date_str}")
    return date_str


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 parse_report.py path/to/Daily_Report_DD-MM-YYYY.xlsx")
        sys.exit(1)
    parse_one(Path(sys.argv[1]))


if __name__ == "__main__":
    main()
